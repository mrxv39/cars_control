#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de importación completa de datos históricos
Importa: STOCK (vehículos), FACTURAS (ventas), CLIENTES
Destino: Carpeta de datos de la app
"""

import openpyxl
from pathlib import Path
import json
import os
from datetime import datetime
from typing import List, Dict, Optional

# Configuración
DOCS_LEGACY = Path("docs_legacy")
APP_DATA_DIR = Path("app/src-tauri/data")
STOCK_DIR = APP_DATA_DIR / "stock"

# Asegurar que existen los directorios
APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
STOCK_DIR.mkdir(parents=True, exist_ok=True)

class DataImporter:
    def __init__(self):
        self.vehicles = []
        self.leads = []  # Leads = clientes potenciales
        self.clients = []  # Clientes confirmados (de facturas)
        self.sales_records = []
        self.errors = []

    def import_stock(self):
        """Importa vehículos del archivo STOCK"""
        print("\n" + "="*80)
        print("IMPORTANDO STOCK (VEHÍCULOS)")
        print("="*80)

        stock_files = [
            "docs_legacy/varios codinacars/STOCK 2025 CODINACARS..xlsx",
            "docs_legacy/varios codinacars/STOCK RICARD CODINA LUDEÑA.xlsx",
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2023/STOCK 2023 RICHI.xlsx",
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2024/2024...varios/STOCK 2024 RICHI.xlsx",
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2025/STOCK 2025 RICHI.xlsx",
        ]

        for file_path in stock_files:
            path = Path(file_path)
            if not path.exists():
                print(f"[WARN] Archivo no encontrado: {file_path}")
                continue

            print(f"\nProcesando: {path.name}")
            try:
                wb = openpyxl.load_workbook(path)

                # Buscar hoja de stock
                sheet_name = None
                for name in wb.sheetnames:
                    if name.upper() == "STOCK":
                        sheet_name = name
                        break

                if not sheet_name:
                    sheet_name = wb.sheetnames[0]  # Primera hoja por defecto

                ws = wb[sheet_name]

                # Encontrar fila de encabezados
                header_row = None
                for row_idx in range(1, 5):
                    row_data = [cell.value for cell in ws[row_idx]]
                    if any(h and 'MARCA' in str(h).upper() for h in row_data):
                        header_row = row_idx
                        break

                if not header_row:
                    header_row = 2

                # Obtener encabezados
                headers = {}
                for col_idx, cell in enumerate(ws[header_row], 1):
                    if cell.value:
                        headers[str(cell.value).upper()] = col_idx

                # Importar filas
                imported = 0
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    # Obtener datos necesarios
                    marca = ws.cell(row=row_idx, column=headers.get("MARCA", 2)).value
                    modelo = ws.cell(row=row_idx, column=headers.get("MODELO", 3)).value

                    if not marca or not modelo:
                        continue

                    año = ws.cell(row=row_idx, column=headers.get("AÑO", 6)).value
                    km = ws.cell(row=row_idx, column=headers.get("KM", 5)).value
                    matricula = ws.cell(row=row_idx, column=headers.get("MATRICULA", 7)).value
                    precio_compra = ws.cell(row=row_idx, column=headers.get("PRECIO", 8)).value
                    pvp = ws.cell(row=row_idx, column=headers.get("PVP", 16)).value

                    # Crear nombre del vehículo
                    name = f"{marca} {modelo}".strip()
                    folder_path = STOCK_DIR / name.replace(" ", "_")

                    # Crear carpeta
                    folder_path.mkdir(parents=True, exist_ok=True)

                    # Crear registro
                    vehicle = {
                        "name": name,
                        "folder_path": str(folder_path),
                        "ad_info": None,
                        "precio_compra": float(precio_compra) if isinstance(precio_compra, (int, float)) else None,
                        "precio_venta": float(pvp) if isinstance(pvp, (int, float)) else None,
                        "km": int(km) if isinstance(km, (int, float)) else None,
                        "anio": int(año) if isinstance(año, (int, float)) else None,
                        "estado": "disponible"
                    }

                    self.vehicles.append(vehicle)
                    imported += 1

                print(f"  [OK] Importados {imported} vehiculos")

            except Exception as e:
                error = f"Error en {path.name}: {str(e)}"
                print(f"  [ERROR] {error}")
                self.errors.append(error)

        print(f"\nTOTAL VEHÍCULOS: {len(self.vehicles)}")

    def import_sales_and_clients(self):
        """Importa facturas de venta y extrae datos de clientes"""
        print("\n" + "="*80)
        print("IMPORTANDO FACTURAS Y CLIENTES")
        print("="*80)

        # Archivos de facturas individuales (contienen datos del cliente)
        factura_files = [
            "docs_legacy/CODINACARS PC/FACTURAS COMPRA - VENTA/FACTURA VENTA RICHI.xlsx",
            "docs_legacy/CODINACARS PC/FACTURAS COMPRA - VENTA/FACTURA VENTA COMISIONES/FACTURA VENTA COMISIONES.xlsx",
        ]

        # Archivos de listado de facturas (para SalesRecords)
        listado_files = [
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2023/LISTADO FACTURAS VENTA 2023 RICHI.xlsx",
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2024/2024...varios/LISTADO FACTURAS VENTA 2024 RICHI.xlsx",
            "docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2025/LISTADO FACTURAS VENTA 2025 RICHI.xlsx",
        ]

        # Procesar facturas individuales para extraer clientes
        print("\nExtrayendo clientes de facturas individuales...")
        for file_path in factura_files:
            path = Path(file_path)
            if not path.exists():
                continue

            print(f"  Procesando: {path.name}")
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active

                # Buscar patrón de datos de cliente
                # Típicamente: Nombre en fila 10, DNI en fila 12-13
                for row_idx in range(1, min(30, ws.max_row)):
                    cell_val = ws.cell(row=row_idx, column=1).value
                    if cell_val and "BERNARDEZ" in str(cell_val).upper():
                        # Encontré un cliente
                        nombre = str(cell_val).strip()

                        # Buscar DNI en filas cercanas
                        dni = None
                        direccion = None
                        ciudad = None

                        for search_row in range(row_idx, min(row_idx + 5, ws.max_row)):
                            search_val = ws.cell(row=search_row, column=1).value
                            if search_val and len(str(search_val)) == 9 and any(c.isdigit() for c in str(search_val)):
                                dni = str(search_val).strip()
                            if search_val and "/" in str(search_val) and "C/" in str(search_val).upper():
                                direccion = str(search_val).strip()

                        # Crear cliente
                        if nombre and dni:
                            client = {
                                "id": len(self.clients) + 1,
                                "name": nombre,
                                "phone": "",
                                "email": "",
                                "dni": dni,
                                "notes": f"Importado de factura {path.name}",
                                "vehicle_folder_path": None,
                                "source_lead_id": None
                            }

                            if client not in self.clients:
                                self.clients.append(client)
                                print(f"    [OK] Cliente: {nombre} ({dni})")
            except Exception as e:
                print(f"  [FAIL] Error: {str(e)}")

        # Procesar listados de facturas para SalesRecords
        print("\nExtrayendo ventas de listados de facturas...")
        for file_path in listado_files:
            path = Path(file_path)
            if not path.exists():
                continue

            print(f"  Procesando: {path.name}")
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active

                # Obtener encabezados
                headers = {}
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value:
                        headers[str(cell.value).upper()] = col_idx

                imported = 0
                for row_idx in range(2, ws.max_row + 1):
                    factura_num = ws.cell(row=row_idx, column=headers.get("N FACTURA", 1)).value
                    fecha = ws.cell(row=row_idx, column=headers.get("FECHA", 2)).value
                    vehiculo = ws.cell(row=row_idx, column=headers.get("VEHICULO", 3)).value
                    matricula = ws.cell(row=row_idx, column=headers.get("MATRICULA", 4)).value
                    base_compra = ws.cell(row=row_idx, column=headers.get("BASE COMPRA", 5)).value
                    neto = ws.cell(row=row_idx, column=headers.get("NETO", 8)).value

                    if not factura_num or not vehiculo:
                        continue

                    # Buscar vehículo en stock
                    vehicle_folder = None
                    for v in self.vehicles:
                        if vehiculo.strip().upper() in v["name"].upper():
                            vehicle_folder = v["folder_path"]
                            break

                    if vehicle_folder:
                        sale = {
                            "id": len(self.sales_records) + 1,
                            "vehicle_folder_path": vehicle_folder,
                            "client_id": None,
                            "lead_id": None,
                            "price_final": float(neto) if isinstance(neto, (int, float)) else 0,
                            "date": str(fecha) if fecha else datetime.now().isoformat(),
                            "notes": f"Factura {factura_num}, Matrícula {matricula}"
                        }
                        self.sales_records.append(sale)
                        imported += 1

                print(f"    [OK] Importadas {imported} ventas")
            except Exception as e:
                print(f"  [FAIL] Error: {str(e)}")

        print(f"\nTOTAL CLIENTES: {len(self.clients)}")
        print(f"TOTAL VENTAS: {len(self.sales_records)}")

    def save_data(self):
        """Guarda los datos en archivos JSON"""
        print("\n" + "="*80)
        print("GUARDANDO DATOS")
        print("="*80)

        # Guardar stock (como carpetas)
        # Ya creadas arriba

        # Guardar clientes
        clients_file = APP_DATA_DIR / "clients.json"
        with open(clients_file, "w", encoding="utf-8") as f:
            json.dump(self.clients, f, ensure_ascii=False, indent=2)
        print(f"[OK] Guardados {len(self.clients)} clientes en {clients_file.name}")

        # Nota: Las ventas se guardarían en SQLite (futuro)
        # Por ahora solo guardamos el análisis

        print(f"\n[OK] IMPORTACIÓN COMPLETADA")
        print(f"  - Vehículos: {len(self.vehicles)}")
        print(f"  - Clientes: {len(self.clients)}")
        print(f"  - Ventas: {len(self.sales_records)}")

        if self.errors:
            print(f"\n[WARN] Errores encontrados: {len(self.errors)}")
            for error in self.errors[:5]:
                print(f"  - {error}")

def main():
    print("\n" + "="*80)
    print("IMPORTADOR DE DATOS HISTÓRICOS - CODINA CARS")
    print("="*80)

    importer = DataImporter()
    importer.import_stock()
    importer.import_sales_and_clients()
    importer.save_data()

    print("\n" + "="*80)
    print("¿Próximos pasos?")
    print("="*80)
    print("""
1. Revisar los datos importados en app/src-tauri/data/
2. Abrir la app: npm run tauri dev
3. Los vehículos deben aparecer en STOCK
4. Los clientes deben aparecer en CLIENTS
5. Las ventas se vincularán en SalesRecords
""")

if __name__ == "__main__":
    main()
