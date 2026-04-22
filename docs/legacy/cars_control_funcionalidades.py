from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1D4ED8')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SECTION_FILL = PatternFill('solid', fgColor='E0E7FF')
SECTION_FONT = Font(name='Arial', bold=True, color='1D4ED8', size=11)
BODY_FONT = Font(name='Arial', size=10)
BODY_FONT_BOLD = Font(name='Arial', size=10, bold=True)
CHECK_FONT = Font(name='Arial', size=12, color='16A34A')
CROSS_FONT = Font(name='Arial', size=12, color='DC2626')
PENDING_FONT = Font(name='Arial', size=10, color='F59E0B')
IMPACT_HIGH = PatternFill('solid', fgColor='FEE2E2')
IMPACT_HIGH_FONT = Font(name='Arial', size=10, bold=True, color='DC2626')
IMPACT_MED = PatternFill('solid', fgColor='FEF3C7')
IMPACT_MED_FONT = Font(name='Arial', size=10, bold=True, color='D97706')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, font=BODY_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = WRAP
        cell.border = thin_border

def section_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical='center')
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = SECTION_FILL

# =====================================================
# SHEET 1: FUNCIONALIDADES ACTIVAS (100% confirmadas)
# =====================================================
ws1 = wb.active
ws1.title = "Activas"
ws1.sheet_properties.tabColor = "16A34A"

headers1 = ["Funcionalidad", "Descripción", "Tauri", "Web", "Módulo / Archivos clave"]
widths1 = [35, 50, 8, 8, 40]
for i, (h, w) in enumerate(zip(headers1, widths1), 1):
    ws1.cell(row=1, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = w
style_header(ws1, 1, 5)

features = [
    # Auth
    ("AUTENTICACIÓN Y USUARIOS", None),
    ("Login usuario/contraseña", "PBKDF2-SHA256 600k iteraciones, migración automática desde SHA-256 legacy", "✓", "✓", "hash.ts, db.rs"),
    ("Roles de usuario", "super_admin, admin, user, readonly", "✓", "✓", "platform-types.ts, db.rs"),
    ("Google OAuth", "Login con cuenta Google vía Supabase Auth (modo Prueba)", "✗", "✓", "platform-api.ts, supabase.ts"),
    ("Persistencia de sesión", "Mantiene sesión entre recargas (localStorage)", "✗", "✓", "WebApp.tsx"),

    # Stock
    ("GESTIÓN DE STOCK / VEHÍCULOS", None),
    ("CRUD vehículos", "Crear, editar (nombre, precios, km, año, estado, combustible, color, notas), eliminar", "✓", "✓", "App.tsx, api.ts, db.rs"),
    ("Estados de vehículo", "Disponible → Reservado → Vendido", "✓", "✓", "StockDetailView.tsx, WebApp.tsx"),
    ("Galería de fotos", "Subir múltiples fotos, eliminar, thumbnail automático", "✓", "✓", "StockDetailView.tsx, WebApp.tsx (VDPhotos)"),
    ("Info de anuncio", "URL, estado y fecha del anuncio en coches.net", "✓", "✓", "vehicle_ads table"),
    ("Documentos/facturas", "Subir PDF/JPG/PNG/WEBP, eliminar, descargar", "✗", "✓", "WebApp.tsx (VDFactura), Supabase Storage"),
    ("Abrir carpeta vehículo", "Abre carpeta del vehículo en explorador del sistema", "✓", "✗", "StockDetailView.tsx, lib.rs"),
    ("3 layouts de detalle", "Layout A (sidebar), B (tabs), C (dashboard) — seleccionable por usuario", "✗", "✓", "WebApp.tsx (VehicleDetail)"),

    # Leads
    ("GESTIÓN DE LEADS (Pre-Compra)", None),
    ("CRUD leads", "Nombre, teléfono, email, interés, canal, estado, fecha contacto, notas", "✓", "✓", "LeadsView.tsx, WebApp.tsx, api.ts, db.rs"),
    ("Estados de lead", "Nuevo, Contactado, Negociando, Cerrado, Perdido, Descartado", "✓", "✓", "shared-types.ts"),
    ("Vincular lead a vehículo", "Asociar lead al vehículo de interés", "✓", "✓", "LeadModal.tsx, WebApp.tsx"),
    ("Notas de lead (timeline)", "Añadir/eliminar notas con timestamp por cada lead", "✓", "✓", "LeadNotesPanel.tsx, api.ts (listLeadNotes)"),
    ("Convertir lead → cliente", "Crea cliente con datos del lead, vincula source_lead_id", "✓", "✓", "App.tsx, WebApp.tsx, api.ts, db.rs"),
    ("Búsqueda leads", "Búsqueda sin acentos por nombre, teléfono, interés", "✓", "✓", "useSearchFilter.ts, WebApp.tsx"),
    ("Parseo email coches.net", "Extrae datos de leads desde emails de notificación de coches.net", "✓", "✗", "Implementación confirmada por usuario"),
    ("Eliminar con confirmación", "Diálogo de confirmación antes de borrar", "✓", "✓", "WebApp.tsx"),

    # Clientes
    ("GESTIÓN DE CLIENTES (Post-Compra)", None),
    ("CRUD clientes", "Nombre, teléfono, email, DNI, notas, vehículo vinculado", "✓", "✓", "ClientsView.tsx, WebApp.tsx, api.ts, db.rs"),
    ("Seguimiento origen", "source_lead_id — saber de qué lead vino el cliente", "✓", "✓", "clients table"),
    ("Edición inline", "Editar datos del cliente sin abrir modal", "✗", "✓", "WebApp.tsx (ClientsList)"),
    ("Eliminar con confirmación", "Diálogo de confirmación antes de borrar", "✓", "✓", "WebApp.tsx"),

    # Ventas
    ("REGISTROS DE VENTAS", None),
    ("Registrar venta", "Vehículo, cliente, lead, precio final, notas, fecha", "✓", "✓", "SalesRecordsView.tsx, WebApp.tsx, api.ts, db.rs"),
    ("Estadísticas ventas", "Total facturado, media por venta, mejor venta, beneficio total", "✓", "✓", "SalesRecordsView.tsx, WebApp.tsx"),
    ("Ordenar ventas", "Por fecha, precio, nombre de vehículo", "✓", "✓", "SalesRecordsView.tsx, WebApp.tsx"),
    ("Informe PDF ventas", "Genera PDF con resumen y tabla de ventas (jsPDF)", "✓", "✗", "reportGenerator.ts"),
    ("Informe mensual PDF", "Filtra ventas por mes y genera informe", "✓", "✗", "reportGenerator.ts"),
    ("Eliminar con confirmación", "Diálogo de confirmación antes de borrar", "✓", "✓", "WebApp.tsx"),

    # Compras/Gastos
    ("COMPRAS Y GASTOS", None),
    ("Registrar gasto", "14 tipos: compra vehículo, taller, transporte, combustible, etc.", "✓", "✓", "PurchasesView.tsx, WebApp.tsx, api.ts, db.rs"),
    ("Filtrar por tipo", "Dropdown para filtrar gastos por categoría", "✓", "✓", "PurchasesView.tsx, WebApp.tsx"),
    ("Ordenar gastos", "Por fecha, precio, proveedor, tipo", "✓", "✓", "PurchasesView.tsx, WebApp.tsx"),
    ("Eliminar con confirmación", "Diálogo de confirmación antes de borrar", "✓", "✓", "WebApp.tsx"),

    # Proveedores
    ("PROVEEDORES", None),
    ("CRUD proveedores (Web)", "Nombre, CIF/NIF, contacto, teléfono, email, notas", "✗", "✓", "WebApp.tsx (SuppliersList), api.ts"),
    ("Vista agregada (Tauri)", "Total gastado, nº facturas, tipos, matrículas — derivado de compras", "✓", "✗", "SuppliersView.tsx"),

    # Dashboard
    ("DASHBOARD", None),
    ("Resumen stock", "Contadores: disponibles, reservados, vendidos, margen potencial", "✓", "✓", "DashboardView.tsx, WebApp.tsx"),
    ("Pipeline leads", "Contadores por estado + leads sin seguimiento", "✓", "✓", "DashboardView.tsx, WebApp.tsx"),
    ("Últimos leads", "Lista de los 5 últimos leads registrados", "✓", "✓", "DashboardView.tsx, WebApp.tsx"),
    ("Gráfico ventas 12 meses", "Barras ventas vs gastos mensual + tabla detalle", "✗", "✓", "WebApp.tsx (WebDashboard)"),
    ("Informe margen", "Top 10 vehículos vendidos con compra, venta, gastos, margen", "✗", "✓", "WebApp.tsx (WebDashboard)"),

    # Recordatorios
    ("RECORDATORIOS / SEGUIMIENTO", None),
    ("Leads sin contacto inicial", "Detecta leads nuevos que no han sido contactados", "✓", "✓", "RemindersView.tsx, WebApp.tsx"),
    ("Leads >7 días sin contacto", "Detecta leads activos sin contacto reciente", "✓", "✓", "RemindersView.tsx, WebApp.tsx"),
    ("Marcar como contactado", "Actualiza fecha_contacto a hoy con un clic", "✓", "✓", "RemindersView.tsx, WebApp.tsx"),

    # Export/Import
    ("EXPORTACIÓN / IMPORTACIÓN", None),
    ("Exportar CSV", "Exporta datos de cualquier lista con formato europeo (;) y BOM", "✗", "✓", "csv-export.ts"),
    ("Backup completo", "Exporta leads, clientes, vehicle_ads en JSON con timestamp", "✓", "✗", "lib.rs (export_app_data)"),
    ("Importar CSV stock", "Importa vehículos desde CSV con mapeo flexible de columnas", "✓", "✗", "importer.rs"),
    ("Búsqueda global", "Busca en vehículos, leads, clientes simultáneamente", "✗", "✓", "WebApp.tsx (GlobalSearchResults)"),

    # Inspecciones
    ("INSPECCIÓN / REVISIÓN TÉCNICA", None),
    ("Ficha inspección", "7 secciones, 30+ items: exterior, interior, motor, frenos, eléctrico, docs", "✗", "✓", "WebApp.tsx (RevisionSheet)"),
    ("Historial inspecciones", "Guarda y muestra inspecciones anteriores por vehículo", "✗", "✓", "vehicle_inspections table"),

    # Catálogo público
    ("CATÁLOGO PÚBLICO", None),
    ("Web pública vehículos", "Listado público de vehículos de CodinaCars con búsqueda y filtros", "✗", "✓", "WebApp.tsx (PublicCatalog)"),
    ("Detalle público", "Ficha completa: fotos, specs, financiación, contacto", "✗", "✓", "WebApp.tsx (PublicVehicleDetail)"),
    ("Formulario contacto", "Envía email a codinacars@gmail.com vía FormSubmit.co", "✗", "✓", "WebApp.tsx (ContactForm)"),
    ("Botón WhatsApp", "Enlace directo a WhatsApp de CodinaCars", "✗", "✓", "WebApp.tsx"),

    # Platform Admin
    ("PLATAFORMA MULTI-TENANT (Super Admin)", None),
    ("Dashboard plataforma", "KPIs globales: empresas, usuarios, vehículos, leads, registros pendientes", "✗", "✓", "PlatformDashboard.tsx"),
    ("Gestión empresas", "Listar, editar, suspender/reactivar, gestionar planes", "✗", "✓", "PlatformCompanies.tsx"),
    ("Revisión registros", "Aprobar/rechazar solicitudes de alta con motivo", "✗", "✓", "PlatformRegistrations.tsx"),
    ("Configuración plataforma", "Ajustes globales clave-valor", "✗", "✓", "PlatformConfig.tsx"),
    ("Registro público empresas", "Formulario de auto-registro para nuevas empresas", "✗", "✓", "RegistrationPage.tsx"),
    ("Gestión usuarios empresa", "Crear usuarios, activar/desactivar", "✗", "✓", "platform-api.ts"),
    ("Invitaciones por email", "Enviar invitaciones con token y expiración", "✗", "✓", "platform-api.ts"),

    # Legacy
    ("DATOS LEGACY (Solo Desktop)", None),
    ("Historial ventas legacy", "Navega carpetas de ventas históricas", "✓", "✗", "SalesView.tsx, lib.rs"),
    ("Fiscal legacy", "Navega carpetas de documentos fiscales", "✓", "✗", "LegacyView.tsx, lib.rs"),
    ("Gastos legacy", "Navega carpetas de gastos históricos", "✓", "✗", "LegacyView.tsx, lib.rs"),

    # Tech
    ("INFRAESTRUCTURA TÉCNICA", None),
    ("78 tests TypeScript", "Suite de tests con Vitest", "✓", "✓", "app/src/test/"),
    ("32 tests Rust", "Tests unitarios del backend", "✓", "✗", "db.rs #[cfg(test)]"),
    ("Error Boundary", "Captura errores de React y muestra fallback amigable", "✓", "✓", "ErrorBoundary.tsx"),
    ("Sugerencias contextuales", "FeedbackButton con tips según vista y datos", "✗", "✓", "FeedbackButton.tsx"),
    ("Paginación", "Hook reutilizable para listas >50 items", "✗", "✓", "usePagination.ts"),
    ("Variables CSS", "Colores centralizados en :root", "✓", "✓", "App.css"),
]

row = 2
for feat in features:
    if feat[1] is None:
        section_row(ws1, row, 5, feat[0])
    else:
        ws1.cell(row=row, column=1, value=feat[0])
        ws1.cell(row=row, column=2, value=feat[1])
        c3 = ws1.cell(row=row, column=3, value=feat[2])
        c4 = ws1.cell(row=row, column=4, value=feat[3])
        ws1.cell(row=row, column=5, value=feat[4])
        c3.alignment = CENTER
        c4.alignment = CENTER
        c3.font = CHECK_FONT if feat[2] == "✓" else CROSS_FONT
        c4.font = CHECK_FONT if feat[3] == "✓" else CROSS_FONT
        style_row(ws1, row, 5)
        c3.font = CHECK_FONT if feat[2] == "✓" else CROSS_FONT
        c4.font = CHECK_FONT if feat[3] == "✓" else CROSS_FONT
        c3.alignment = CENTER
        c4.alignment = CENTER
    row += 1

ws1.auto_filter.ref = f"A1:E{row-1}"
ws1.freeze_panes = 'A2'

# =====================================================
# SHEET 2: PLANIFICADAS (del PLAN_MEJORAS.md)
# =====================================================
ws2 = wb.create_sheet("Planificadas")
ws2.sheet_properties.tabColor = "F59E0B"

headers2 = ["Fase", "Tarea", "Descripción", "Estado", "Riesgo", "Archivos"]
widths2 = [12, 35, 50, 14, 10, 35]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    ws2.cell(row=1, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header(ws2, 1, 6)

DONE_FILL = PatternFill('solid', fgColor='DCFCE7')
SKIP_FILL = PatternFill('solid', fgColor='FEF3C7')

planned = [
    ("Fase 1", "1.1 Corregir 4 errores TS", "Eliminar imports y variables no usadas", "✅ Completada", "Bajo", "StockDetailView, tests"),
    ("Fase 1", "1.2 CSS duplicado", "Fusionar regla .record-notes duplicada", "✅ Completada", "Bajo", "App.css"),
    ("Fase 1", "1.3 Variables CSS", "Extraer colores hardcoded a :root variables", "✅ Completada", "Bajo", "App.css"),
    ("Fase 2", "2.1 Tests hash.ts", "8 tests unitarios para PBKDF2 y legacy SHA-256", "✅ Completada", "Nulo", "hash.test.ts"),
    ("Fase 2", "2.2 Tests shared-types", "Tests para tipos y constantes compartidas", "✅ Completada", "Nulo", "shared-types.test.ts"),
    ("Fase 2", "2.3 Tests api.ts", "6 tests para capa Supabase (mockeados)", "✅ Completada", "Nulo", "api.test.ts"),
    ("Fase 2", "2.4 Tests platform-api", "4 tests para API multi-tenant", "✅ Completada", "Nulo", "platform-api.test.ts"),
    ("Fase 3", "3.1 Dashboard web métricas", "Stock, leads, margen, ventas mes, alertas", "✅ Completada", "Bajo", "WebApp.tsx"),
    ("Fase 3", "3.2 Exportar CSV", "Botón CSV en stock, leads, clientes, ventas, compras, proveedores", "✅ Completada", "Bajo", "csv-export.ts, WebApp.tsx"),
    ("Fase 3", "3.3 Búsqueda global", "Buscar en vehículos, leads, clientes desde barra superior", "✅ Completada", "Bajo", "WebApp.tsx"),
    ("Fase 3", "3.4 Recordatorios web", "Vista de leads sin contactar / >7 días sin contacto", "✅ Completada", "Bajo", "WebApp.tsx"),
    ("Fase 3", "3.5 Edición leads/clientes web", "Edición inline de datos de leads y clientes", "✅ Completada", "Medio", "WebApp.tsx"),
    ("Fase 3", "3.6 Lead → cliente web", "Conversión de lead a cliente desde la web", "✅ Completada", "Medio", "WebApp.tsx"),
    ("Fase 4", "4.1 Extraer componentes WebApp", "Mover componentes inline a archivos separados (~1778 líneas)", "⏭ Omitida", "Medio", "WebApp.tsx → components/web/"),
    ("Fase 4", "4.2 Error boundaries", "Captura de errores de React con fallback amigable", "✅ Completada", "Bajo", "ErrorBoundary.tsx"),
    ("Fase 4", "4.3 Accesibilidad aria", "aria-label, role=status, role=alert en componentes", "✅ Completada", "Bajo", "Varios componentes"),
    ("Fase 4", "4.4 Mensajes de error", "Mensajes descriptivos en lugar de String(err)", "✅ Completada", "Bajo", "WebApp.tsx, api.ts"),
    ("Fase 5", "5.1 Memoización", "useMemo y useCallback en cálculos costosos", "✅ Completada", "Bajo", "WebApp.tsx"),
    ("Fase 5", "5.2 Paginación", "Hook usePagination para listas >50 items", "✅ Completada", "Medio", "usePagination.ts, WebApp.tsx"),
    ("Fase 5", "5.3 Lazy loading fotos", "loading='lazy' en imágenes de vehículos", "✅ Completada", "Bajo", "WebApp.tsx"),
    ("Fase 6", "6.1 Media queries", "Responsive básico para <768px", "✅ Completada", "Bajo", "App.css"),
    ("Fase 6", "6.2 Menú hamburguesa", "Menú desplegable en móvil sustituyendo sidebar", "✅ Completada", "Medio", "WebApp.tsx, App.css"),
    ("Fase 7", "7.1 Inspección ITV", "Persistir fichas de revisión en Supabase", "✅ Completada", "Medio", "RevisionSheet, api.ts"),
    ("Fase 7", "7.2 Notas leads web", "Panel de notas en leads desde la web", "✅ Completada", "Bajo", "WebApp.tsx, api.ts"),
    ("Fase 7", "7.3 Informe margen", "Margen bruto por vehículo vendido con gastos asociados", "✅ Completada", "Bajo", "WebApp.tsx (Dashboard)"),
    ("Fase 7", "7.4 Informe mensual", "Ventas y gastos agrupados por mes con gráfico", "✅ Completada", "Bajo", "WebApp.tsx (Dashboard)"),
    ("Fase 8", "8.1 Tipos error Rust", "AppError con thiserror en lugar de Result<T, String>", "✅ Completada", "Medio", "db.rs, lib.rs"),
    ("Fase 8", "8.2 Transacciones SQL", "Envolver operaciones multi-tabla en transacciones", "✅ Completada", "Medio", "db.rs"),
    ("Fase 8", "8.3 Tests Rust", "Tests unitarios para db.rs (hash, auth, CRUD)", "✅ Completada", "Nulo", "db.rs #[cfg(test)]"),
]

for i, p in enumerate(planned, 2):
    for j, val in enumerate(p, 1):
        ws2.cell(row=i, column=j, value=val)
    style_row(ws2, i, 6)
    if "Completada" in p[3]:
        for c in range(1, 7):
            ws2.cell(row=i, column=c).fill = DONE_FILL
    elif "Omitida" in p[3]:
        for c in range(1, 7):
            ws2.cell(row=i, column=c).fill = SKIP_FILL

ws2.auto_filter.ref = f"A1:F{len(planned)+1}"
ws2.freeze_panes = 'A2'

# =====================================================
# SHEET 3: SUGERENCIAS DE ALTO IMPACTO
# =====================================================
ws3 = wb.create_sheet("Alto Impacto")
ws3.sheet_properties.tabColor = "DC2626"

headers3 = ["#", "Funcionalidad", "Descripción", "Impacto Negocio", "Esfuerzo", "Dependencias", "ROI"]
widths3 = [4, 30, 55, 45, 12, 30, 10]
for i, (h, w) in enumerate(zip(headers3, widths3), 1):
    ws3.cell(row=1, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header(ws3, 1, 7)

suggestions = [
    (1, "Auto-respuesta leads coches.net",
     "Reply automático al email relay de coches.net (@contactos.coches.net) cuando llega un lead. Mensaje personalizado con nombre del cliente y vehículo de interés.",
     "CRÍTICO: coches.net penaliza reputación si no respondes rápido. Ahora depende de que Ricard vea el email y conteste manualmente. Automatizar = 0s tiempo de respuesta.",
     "Medio",
     "Gmail API (proyecto gmail-push-fly ya existe). Verificar que reply al relay funciona (prueba pendiente).",
     "MUY ALTO"),

    (2, "Importar CSV leads de coches.net",
     "Función para importar el CSV que se descarga desde el panel pro de coches.net. Mapear campos automáticamente y crear leads en Cars Control.",
     "Evita perder leads y elimina el trabajo manual de copiar datos uno a uno del panel de coches.net a la app.",
     "Bajo",
     "Ninguna. Exportación CSV confirmada por Elvira (coches.net). Solo falta mapear campos.",
     "ALTO"),

    (3, "WhatsApp Business API para leads",
     "Enviar mensaje de WhatsApp automático al teléfono del lead cuando se crea. Usar WhatsApp Business API (Twilio/360dialog) o enlace wa.me como paso intermedio.",
     "Los clientes de coches responden mucho más rápido por WhatsApp que por email. Reduce tiempo de primer contacto de horas a segundos.",
     "Medio-Alto",
     "WhatsApp Business API requiere cuenta verificada y tiene coste (~0.05€/msg). Alternativa gratuita: generar enlace wa.me para que Ricard haga 1 clic.",
     "ALTO"),

    (4, "Alertas en tiempo real (push/email)",
     "Notificación instantánea cuando llega un lead nuevo: push notification en móvil, email, o Telegram bot. Ricard no necesita estar mirando la app.",
     "Actualmente si Ricard no mira el email o la app, el lead se queda sin atender. Con alertas, el tiempo de respuesta baja drásticamente.",
     "Bajo-Medio",
     "Supabase Edge Functions + webhook. O Telegram Bot API (gratuito, setup en 10 min).",
     "ALTO"),

    (5, "Cálculo automático de rentabilidad",
     "Al registrar una venta: calcular automáticamente beneficio neto = precio_venta - precio_compra - TODOS los gastos asociados (taller, transporte, limpieza, ITV, etc.)",
     "Ricard necesita saber cuánto gana realmente por coche. Ahora tiene que calcular manualmente sumando facturas. Errores frecuentes.",
     "Bajo",
     "Datos ya existen (purchase_records con vehicle_id). Solo falta la query que sume gastos por vehículo.",
     "MUY ALTO"),

    (6, "Pipeline visual de leads (Kanban)",
     "Vista tipo Kanban donde cada columna es un estado (Nuevo → Contactado → Negociando → Cerrado/Perdido). Drag & drop para mover leads entre estados.",
     "Ricard ve de un vistazo dónde está cada lead en el proceso. Más intuitivo que una lista. Reduce leads olvidados.",
     "Medio",
     "Ninguna. Datos y estados ya existen. Solo es UI nueva.",
     "ALTO"),

    (7, "Historial de precios coches.net",
     "Cada vez que se actualiza el precio de un vehículo en la app, guardar log con fecha y precio anterior. Mostrar historial de cambios de precio.",
     "Permite analizar cuánto tiempo tarda en venderse un coche según el precio, y si las bajadas de precio funcionan.",
     "Bajo",
     "Nueva tabla price_history (vehicle_id, old_price, new_price, date). Trigger en updateVehicle.",
     "MEDIO"),

    (8, "Firma digital en inspecciones",
     "Añadir campo de firma digital (canvas) en la ficha de inspección. Guardar como imagen en Supabase Storage.",
     "Da validez profesional a las inspecciones. Útil si hay disputas post-venta sobre el estado del vehículo.",
     "Medio",
     "Librería signature_pad (JS). Guardar como PNG en Supabase Storage.",
     "MEDIO"),

    (9, "App móvil (PWA)",
     "Convertir la web en Progressive Web App instalable. Añadir manifest.json, service worker, iconos. Ricard puede 'instalar' la app en su móvil.",
     "Ricard puede gestionar leads desde el móvil en el taller sin necesitar el PC. Acceso rápido con icono en pantalla.",
     "Bajo",
     "Ya es SPA en Vercel. Solo falta manifest.json + service worker básico + iconos.",
     "ALTO"),

    (10, "Integración email bidireccional",
     "No solo parsear emails entrantes, sino ver el historial completo de emails con cada lead dentro de Cars Control. Gmail API para leer el hilo de conversación.",
     "Ricard ve todo el contexto de comunicación con un lead sin salir de la app. No pierde emails ni contexto.",
     "Alto",
     "Gmail API (lectura). Proyecto gmail-push-fly ya tiene OAuth configurado.",
     "MEDIO"),
]

for i, s in enumerate(suggestions, 2):
    for j, val in enumerate(s, 1):
        ws3.cell(row=i, column=j, value=val)
    style_row(ws3, i, 7)
    ws3.cell(row=i, column=1).alignment = CENTER
    roi = s[6]
    roi_cell = ws3.cell(row=i, column=7)
    roi_cell.alignment = CENTER
    if roi == "MUY ALTO":
        roi_cell.font = IMPACT_HIGH_FONT
        roi_cell.fill = IMPACT_HIGH
    elif roi == "ALTO":
        roi_cell.font = IMPACT_MED_FONT
        roi_cell.fill = IMPACT_MED

ws3.auto_filter.ref = f"A1:G{len(suggestions)+1}"
ws3.freeze_panes = 'A2'

# Row heights for readability
for ws in [ws1, ws2, ws3]:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 28

output = '/sessions/gallant-nice-heisenberg/mnt/cars_control/cars_control_funcionalidades.xlsx'
wb.save(output)
print(f"Saved to {output}")
