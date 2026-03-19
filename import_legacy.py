# -*- coding: utf-8 -*-
"""
Import legacy sales data into Cars Control SQLite database.
Sources:
  - Excel invoice lists (2023, 2024) for exact financial data
  - Folder structure (2025, 2026) for additional sales
"""

import sqlite3
import openpyxl
import datetime
import os
import re

DB_PATH = os.path.expandvars(r'%APPDATA%\com.codinacars.carscontrol\data\app.db')

EXCEL_FILES = [
    r'C:/Users/Usuario/Desktop/proyectos/cars_control/docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2023/LISTADO FACTURAS VENTA 2023 RICHI.xlsx',
    r'C:/Users/Usuario/Desktop/proyectos/cars_control/docs_legacy/varios codinacars/VARIOS, 23,24,25/varios 2024/2024...varios/LISTADO FACTURAS VENTA 2024 RICHI.xlsx',
]

VENTAS_FOLDERS = {
    2025: r'C:/Users/Usuario/Desktop/proyectos/cars_control/docs_legacy/varios codinacars/VENTAS 2025',
    2026: r'C:/Users/Usuario/Desktop/proyectos/cars_control/docs_legacy/CODINACARS PC/VENTAS 2026',
}

VENTAS_FOLDERS_WITH_EXCEL = {
    2024: r'C:/Users/Usuario/Desktop/proyectos/cars_control/docs_legacy/varios codinacars/VENTAS 2024',
}

MONTH_MAP = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
    'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
    'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12,
}

def parse_month_from_folder(folder_name):
    """Extract month number from folder name like '01 ENERO' or '1. ENERO'."""
    folder_upper = folder_name.upper()
    for name, num in MONTH_MAP.items():
        if name in folder_upper:
            return num
    m = re.match(r'(\d+)', folder_name)
    if m:
        return int(m.group(1))
    return None

def extract_plate(text):
    """Try to extract a Spanish plate from text like '3267GGR' or '2984LVF'."""
    m = re.search(r'\b(\d{4}\s*[A-Z]{3})\b', text.upper())
    if m:
        return m.group(1).replace(' ', '')
    m = re.search(r'\b([A-Z]\d{4}[A-Z]{2})\b', text.upper())
    if m:
        return m.group(1)
    return None

def extract_client_from_folder(folder_name):
    """Try to extract client name from folder name like 'CUPRA FORMENTOR 3445 MGK( vasile )'."""
    m = re.search(r'\(\s*([^)]+?)\s*\)', folder_name)
    if m:
        name = m.group(1).strip()
        # Filter out non-name annotations
        skip = ['stock conjunto', 'fotos', 'fra ', 'probar', 'itv', 'cn pasado', 'caixa', 'ok']
        if any(s in name.lower() for s in skip):
            return None
        if len(name) > 2:
            return name.title()
    # Check for names after comma like "PEUGEOT 307 , LUIS SANCHEZ"
    parts = folder_name.split(',')
    if len(parts) >= 2:
        after_comma = parts[-1].strip()
        # Remove plate patterns and annotations
        after_comma = re.sub(r'\d{4}\s*[A-Z]{3}', '', after_comma).strip()
        after_comma = re.sub(r'(FRA\s*\d+|OK|CN PASADO|PASADO)', '', after_comma, flags=re.IGNORECASE).strip()
        if after_comma and len(after_comma) > 3 and after_comma[0].isalpha():
            return after_comma.title()
    return None

def clean_vehicle_name(folder_name):
    """Clean vehicle name from folder: remove plate, annotations, parens."""
    name = folder_name
    # Remove parenthetical notes
    name = re.sub(r'\([^)]*\)', '', name)
    # Remove plate
    name = re.sub(r'\b\d{4}\s*[A-Z]{3}\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b[A-Z]\d{4}[A-Z]{2}\b', '', name, flags=re.IGNORECASE)
    # Remove common annotations
    name = re.sub(r'(\.?\s*(OK|CN PASADO|FOTOS OK CC|itv revisar|fra\s*\d+))', '', name, flags=re.IGNORECASE)
    # Remove leading numbering like "3º T ." or "4ºT."
    name = re.sub(r'^\d+[º°]?\s*T\.?\s*\.?\s*', '', name, flags=re.IGNORECASE)
    # Clean up
    name = re.sub(r'[,.\s]+$', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name if name else folder_name


def read_excel_sales(filepath):
    """Read sales from an Excel invoice list."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Hoja1']
    sales = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        factura, fecha, vehiculo, matricula, base_compra, b_imponible, beneficio, neto, iva = row[:9]
        if not vehiculo or not fecha:
            continue
        if isinstance(fecha, datetime.datetime):
            date_str = fecha.strftime('%Y-%m-%d')
            month = fecha.month
            year = fecha.year
        else:
            # Parse string date dd/mm/yyyy
            parts = str(fecha).split('/')
            if len(parts) == 3:
                date_str = f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
                month = int(parts[1])
                year = int(parts[2])
            else:
                continue

        sales.append({
            'invoice': str(factura) if factura else None,
            'date': date_str,
            'month': month,
            'year': year,
            'vehicle': str(vehiculo).strip(),
            'plate': str(matricula).strip() if matricula else None,
            'purchase_price': float(base_compra) if base_compra else None,
            'sale_price': float(b_imponible) if b_imponible else 0,
            'profit': float(beneficio) if beneficio else None,
        })
    return sales


def read_folder_sales(year, base_path):
    """Read sales from folder structure."""
    sales = []
    if not os.path.exists(base_path):
        return sales
    for month_folder in sorted(os.listdir(base_path)):
        month_path = os.path.join(base_path, month_folder)
        if not os.path.isdir(month_path):
            continue
        month_num = parse_month_from_folder(month_folder)
        for vehicle_folder in sorted(os.listdir(month_path)):
            vehicle_path = os.path.join(month_path, vehicle_folder)
            if not os.path.isdir(vehicle_path):
                continue
            plate = extract_plate(vehicle_folder)
            client = extract_client_from_folder(vehicle_folder)
            vehicle_name = clean_vehicle_name(vehicle_folder)
            date_str = f'{year}-{month_num:02d}-15' if month_num else f'{year}-01-15'

            sales.append({
                'invoice': None,
                'date': date_str,
                'month': month_num,
                'year': year,
                'vehicle': vehicle_name,
                'plate': plate,
                'purchase_price': None,
                'sale_price': 0,
                'profit': None,
                'client_name': client,
                'folder_path': vehicle_path,
            })
    return sales


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")  # Temporarily disable for import

    now = datetime.datetime.utcnow().isoformat() + 'Z'
    clients_added = {}
    vehicles_added = set()
    sales_count = 0

    # 1. Read Excel sales (2023, 2024) - these have exact financial data
    excel_sales = []
    for filepath in EXCEL_FILES:
        excel_sales.extend(read_excel_sales(filepath))

    # Build a set of plates from Excel for dedup with folder sales
    excel_plates = {s['plate'] for s in excel_sales if s['plate']}

    # 2. Read folder sales for 2024 (to get client names), match with Excel data
    folder_2024 = read_folder_sales(2024, VENTAS_FOLDERS_WITH_EXCEL[2024])
    folder_2024_by_plate = {}
    for fs in folder_2024:
        if fs.get('plate'):
            folder_2024_by_plate[fs['plate']] = fs

    # Enrich Excel 2024 sales with client names from folders
    for sale in excel_sales:
        if sale['year'] == 2024 and sale.get('plate') and sale['plate'] in folder_2024_by_plate:
            folder_info = folder_2024_by_plate[sale['plate']]
            sale['client_name'] = folder_info.get('client_name')
            sale['folder_path'] = folder_info.get('folder_path')
        else:
            sale.setdefault('client_name', None)
            sale.setdefault('folder_path', None)

    # 3. Read folder sales for 2025 and 2026
    folder_sales = []
    for year, base_path in VENTAS_FOLDERS.items():
        folder_sales.extend(read_folder_sales(year, base_path))

    # Combine all sales
    all_sales = excel_sales + folder_sales

    print(f"Total sales to import: {len(all_sales)}")
    print(f"  From Excel: {len(excel_sales)}")
    print(f"  From folders (2025+2026): {len(folder_sales)}")
    print()

    # 4. Insert into database
    for sale in all_sales:
        # Create vehicle entry if we have plate
        vehicle_key = sale.get('plate') or sale['vehicle']
        if vehicle_key not in vehicles_added:
            folder_path = sale.get('folder_path') or f"legacy/{sale['year']}/{vehicle_key}"
            conn.execute(
                "INSERT OR IGNORE INTO vehicles (folder_path, name, precio_compra, precio_venta, estado) VALUES (?, ?, ?, ?, ?)",
                (folder_path, sale['vehicle'], sale.get('purchase_price'), sale.get('sale_price'), 'vendido')
            )
            vehicles_added.add(vehicle_key)

        # Create client if we have a name
        client_id = None
        client_name = sale.get('client_name')
        if client_name and client_name not in clients_added:
            cursor = conn.execute(
                "INSERT INTO clients (name, phone, email, dni, notes, vehicle_folder_path) VALUES (?, '', '', '', ?, NULL)",
                (client_name, f"Importado de legacy - {sale['vehicle']}")
            )
            client_id = cursor.lastrowid
            clients_added[client_name] = client_id
            print(f"  Client: {client_name} (id={client_id})")
        elif client_name:
            client_id = clients_added[client_name]

        # Insert sales_record
        vehicle_folder = sale.get('folder_path') or f"legacy/{sale['year']}/{vehicle_key}"
        notes_parts = []
        if sale.get('invoice'):
            notes_parts.append(f"Factura: {sale['invoice']}")
        if sale.get('profit') and sale['profit'] > 0:
            notes_parts.append(f"Beneficio: {sale['profit']:.0f} EUR")
        if sale.get('purchase_price') and sale['purchase_price'] > 0:
            notes_parts.append(f"Compra: {sale['purchase_price']:.0f} EUR")
        if sale.get('plate'):
            notes_parts.append(f"Matricula: {sale['plate']}")
        notes = ' | '.join(notes_parts) if notes_parts else ''

        conn.execute(
            "INSERT INTO sales_records (vehicle_folder_path, client_id, lead_id, price_final, date, notes) VALUES (?, ?, NULL, ?, ?, ?)",
            (vehicle_folder, client_id, sale.get('sale_price', 0), sale['date'] + 'T12:00:00Z', notes)
        )
        sales_count += 1

    conn.commit()

    # Print summary
    print(f"\n=== Import Complete ===")
    print(f"Sales records: {sales_count}")
    print(f"Vehicles: {len(vehicles_added)}")
    print(f"Clients: {len(clients_added)}")

    # Verify
    count = conn.execute("SELECT COUNT(*) FROM sales_records").fetchone()[0]
    print(f"\nVerification - sales_records in DB: {count}")
    count = conn.execute("SELECT COUNT(*) FROM vehicles").fetchone()[0]
    print(f"Verification - vehicles in DB: {count}")
    count = conn.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    print(f"Verification - clients in DB: {count}")

    conn.close()


if __name__ == '__main__':
    main()
